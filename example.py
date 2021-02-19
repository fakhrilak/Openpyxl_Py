from openpyxl import Workbook,load_workbook
wb = Workbook()
ws = wb.active
#menulis menggunakan perulangan
for a in range(80):
    A = str("A"+str(a+1))
    C = str("C"+str(a+1))
    #menulis di kolom A
    ws[A] = a
    #menulis di kolom B
    ws[C] = "Hello world "+ str(a)
wb.save("sample.xlsx")

obj = load_workbook('sample.xlsx')
sheet= obj.active

#membaca dengan perulangan
for a in range (80):
    obj_cell = sheet.cell(row=a+1,column =3)
    print(obj_cell.value)